from collections import defaultdict
import csv
from datetime import timedelta
import io
import json
import os
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from django.shortcuts import get_object_or_404
import xlwt
from .models import Enquiry, FollowUp_History, Retail_FollowUp_History, clients, confirmed_enquiry, personal_Accident, retailpolicies


def export_confirmed_orders_csv(request):
    # Get confirmed orders with related confirmed_enquiry data
    confirmed_orders = Enquiry.objects.filter(is_confirmed=True).select_related('created_by').prefetch_related('confirmed_enquiry_set')

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Active_policies.csv"'

    writer = csv.writer(response)

    # Write header row
    writer.writerow(['S.No', 'Company Name', 'Customer Name', 'Contact', 'Policies', 'Start Date', 'End Date', 'Policy Number'])

    # Write data rows
    index = 1
    for order in confirmed_orders:
        confirmed_entries = order.confirmed_enquiry_set.all()
        if confirmed_entries.exists():
            for confirmed_entry in confirmed_entries:
                writer.writerow([
                    index,
                    order.companyname or "N/A",
                    order.customername or "N/A",
                    order.contact or "N/A",
                    order.products or "N/A",
                    confirmed_entry.startdate.strftime('%Y-%m-%d') if confirmed_entry.startdate else "N/A",
                    confirmed_entry.enddate.strftime('%Y-%m-%d') if confirmed_entry.enddate else "N/A",
                    confirmed_entry.vid if not confirmed_entry.relegate else "Relegated",
                ])
                index += 1
        else:
            writer.writerow([
                index,
                order.companyname or "N/A",
                order.customername or "N/A",
                order.contact or "N/A",
                order.products or "N/A",
                "N/A", "N/A", "No confirmed entry"
            ])
            index += 1

    return response


from django.http import HttpResponse
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Enquiry
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Enquiry

def export_confirmed_orders_xlsx(request):
    # Fetch data
    current_user = request.user
    if current_user.is_superuser:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True).prefetch_related('confirmed_enquiry_set', 'products')
    else:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True, created_by=current_user).prefetch_related('confirmed_enquiry_set', 'products')

    # Create an Excel workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Active Policies"

    # Write header row
    headers = ["S.No", "Company Name", "Customer Name", "Contact", "Policies", "Start Date", "End Date", "Policy Number"]
    worksheet.append(headers)

    # Write data rows
    index = 1
    for order in confirmed_orders:
        confirmed_entries = order.confirmed_enquiry_set.all()
        product_name = order.products.name if order.products else "N/A"  # ✅ Fixed ForeignKey Handling

        if confirmed_entries.exists():
            for confirmed_entry in confirmed_entries:
                row = [
                    index,
                    order.companyname or "N/A",
                    order.customername or "N/A",
                    str(order.contact) if order.contact else "N/A",
                    product_name,  # ✅ Corrected ForeignKey Handling
                    confirmed_entry.startdate.strftime('%Y-%m-%d') if confirmed_entry.startdate else "N/A",
                    confirmed_entry.enddate.strftime('%Y-%m-%d') if confirmed_entry.enddate else "N/A",
                    confirmed_entry.vid if not confirmed_entry.relegate else "Relegated",
                ]
                worksheet.append(row)
                index += 1
        else:
            row = [
                index,
                order.companyname or "N/A",
                order.customername or "N/A",
                str(order.contact) if order.contact else "N/A",
                product_name,
                "N/A",  # No Start Date
                "N/A",  # No End Date
                "No Confirmed Entry",
            ]
            worksheet.append(row)
            index += 1

    # Prepare the HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="confirmed_orders.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response


from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import Enquiry

def export_confirmed_orders_pdf(request):
    # Fetch data
    current_user = request.user
    if current_user.is_superuser:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True).prefetch_related('confirmed_enquiry_set', 'products')
    else:
        confirmed_orders = Enquiry.objects.filter(is_confirmed=True, created_by=current_user).prefetch_related('confirmed_enquiry_set', 'products')

    # Create a PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Active_policies.pdf"'

    # Set up PDF canvas
    pdf_canvas = canvas.Canvas(response, pagesize=letter)
    pdf_canvas.setFont("Helvetica", 12)

    y = 750  # Initial Y position for content

    # Headers and Column Positions
    headers = ["S.No", "Company", "Customer", "Contact", "Policies", "Start Date", "End Date", "Policy Number"]
    column_positions = [50, 100, 230, 330, 430, 530, 630, 730]  # Adjust spacing for columns

    def draw_headers():
        """Function to draw table headers."""
        nonlocal y
        pdf_canvas.setFont("Helvetica-Bold", 10)
        for i, header in enumerate(headers):
            pdf_canvas.drawString(column_positions[i], y, header)
        y -= 20  # Move down after headers
        pdf_canvas.setFont("Helvetica", 10)  # Reset font

    # Title
    pdf_canvas.setFont("Helvetica-Bold", 14)
    pdf_canvas.drawString(50, y, "Active Policies")
    y -= 20
    draw_headers()

    # Write data rows
    index = 1

    for order in confirmed_orders:
        confirmed_entries = order.confirmed_enquiry_set.all()
        product_name = order.products.name if order.products else "N/A"  # ✅ Fixed ForeignKey Handling

        if confirmed_entries.exists():
            for confirmed_entry in confirmed_entries:
                pdf_canvas.drawString(column_positions[0], y, str(index))  # Row number
                pdf_canvas.drawString(column_positions[1], y, order.companyname or "N/A")  # Company Name
                pdf_canvas.drawString(column_positions[2], y, order.customername or "N/A")  # Customer Name
                pdf_canvas.drawString(column_positions[3], y, str(order.contact) if order.contact else "N/A")  # Contact
                pdf_canvas.drawString(column_positions[4], y, product_name)  # ✅ Fixed ForeignKey Handling
                pdf_canvas.drawString(column_positions[5], y, confirmed_entry.startdate.strftime('%Y-%m-%d') if confirmed_entry.startdate else "N/A")  # ✅ Start Date
                pdf_canvas.drawString(column_positions[6], y, confirmed_entry.enddate.strftime('%Y-%m-%d') if confirmed_entry.enddate else "N/A")  # ✅ End Date
                pdf_canvas.drawString(column_positions[7], y, confirmed_entry.vid if not confirmed_entry.relegate else "Relegated")  # ✅ VID/Relegated

                y -= 20
                index += 1

                # Page Break Handling
                if y < 50:  
                    pdf_canvas.showPage()
                    y = 750
                    draw_headers()  # Redraw headers on new page

        else:
            pdf_canvas.drawString(column_positions[0], y, str(index))
            pdf_canvas.drawString(column_positions[1], y, order.companyname or "N/A")
            pdf_canvas.drawString(column_positions[2], y, order.customername or "N/A")
            pdf_canvas.drawString(column_positions[3], y, str(order.contact) if order.contact else "N/A")
            pdf_canvas.drawString(column_positions[4], y, product_name)
            pdf_canvas.drawString(column_positions[5], y, "N/A")  # Start Date
            pdf_canvas.drawString(column_positions[6], y, "N/A")  # End Date
            pdf_canvas.drawString(column_positions[7], y, "N/A")  # VID/Relegated

            y -= 20
            index += 1

            # Page Break Handling
            if y < 50:  
                pdf_canvas.showPage()
                y = 750
                draw_headers()  # Redraw headers on new page

    pdf_canvas.save()
    return response


def export_lost_enquiries_csv(request):
    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.csv"'

    writer = csv.writer(response)
    writer.writerow([ 'Company Name', 'Customer Name', 'Contact', 'Policies', 'Closing Date', 'Remarks'])

    for enquiry in lost_enquiries:
        writer.writerow([  enquiry.companyname, enquiry.customername,  enquiry.contact, enquiry.products.name,enquiry.closuredate, enquiry.remarks,])

    return response


def export_lost_enquiries_xlsx(request):
    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lost Enquiries"

    # Write header row
    ws.append([ 'Company Name', 'Customer Name', 'Contact', 'Policies', 'Closing Date', 'Remarks'])

    for enquiry in combined_enquiries:
        ws.append([  enquiry.companyname, enquiry.customername,  enquiry.contact, enquiry.products.name,enquiry.closuredate, enquiry.remarks,])

    # Create an in-memory file and save the workbook
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Return the response
    response = HttpResponse(stream, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.xlsx"'

    return response

def export_lost_enquiries_pdf(request):
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    current_user = request.user
    if current_user.is_superuser:
        lost_enquiries = Enquiry.objects.filter(is_lost=True)
        relegated_enquiries = Enquiry.objects.filter(is_relegated=True)
    else:
        lost_enquiries = Enquiry.objects.filter(is_lost=True, created_by=current_user)
        relegated_enquiries = Enquiry.objects.filter(
            is_relegated=True,
            created_by=current_user
        )
    combined_enquiries = lost_enquiries.union(relegated_enquiries)
    print(combined_enquiries)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lost_enquiries.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Define some fonts and styles
    p.setFont("Helvetica-Bold", 10)

    # Table header with corrected alignment and spacing
    p.drawString(50, height - 40, "Company Name")
    p.drawString(150, height - 40, "Customer Name")
    p.drawString(250, height - 40, "Contact")
    p.drawString(350, height - 40, "Policies")
    p.drawString(450, height - 40, "Closing Date")
    p.drawString(550, height - 40, "Remarks")

    y_position = height - 60  # Adjust starting position

    p.setFont("Helvetica", 10)  # Set normal font for data rows

    for enquiry in combined_enquiries:
        p.drawString(50, y_position, str(enquiry.companyname)[:25])  # Truncate long names
        p.drawString(150, y_position, str(enquiry.customername)[:25])
        p.drawString(250, y_position, str(enquiry.contact))
        p.drawString(350, y_position, str(enquiry.products.name) if enquiry.products else "N/A")
        p.drawString(450, y_position, enquiry.closuredate.strftime("%Y-%m-%d") if enquiry.closuredate else "N/A")
        p.drawString(550, y_position, str(enquiry.remarks)[:40])  # Limit remarks to avoid overflow

        y_position -= 20  # Move to the next row

        if y_position < 50:  # Handle page break
            p.showPage()
            y_position = height - 60
            p.setFont("Helvetica", 10)

    p.showPage()
    p.save()

    # Save to response
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response








# ---------------------------------------------------------------------------------------------------------------
import csv
from django.http import HttpResponse
from django.db.models import Q
from .models import Enquiry, User

def export_enquiries_csv(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    # If a specific user is selected
    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()

        # Managers and superusers can see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see only their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="enquiries.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Company Name', 'Customer Name', 'Contact', 'Status', 'Policies', 'Enquiry Date', 'Remarks'])
    
    # Write data to CSV
    for enquiry in enquiries:
        writer.writerow([
            enquiry.id,
            enquiry.companyname,
            enquiry.customername,
            enquiry.contact,
            enquiry.get_status_display(),
            enquiry.products.name,
            enquiry.closuredate,
            enquiry.remarks
        ])
    
    return response


from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.db.models import Q
from .models import Enquiry, User

def export_enquiries_pdf(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()
        # Managers/superusers see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="enquiries.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Table headers
    p.setFont("Helvetica", 10)
    p.drawString(30, height - 40, "Company Name")
    p.drawString(120, height - 40, "Customer Name")
    p.drawString(200, height - 40, "Contact")
    p.drawString(280, height - 40, "Status")
    p.drawString(350, height - 40, "Product")
    p.drawString(450, height - 40, "Enquiry Date")
    p.drawString(550, height - 40, "Remarks")

    y_position = height - 60
    for enquiry in enquiries:
        p.drawString(30, y_position, str(enquiry.companyname))
        p.drawString(120, y_position, str(enquiry.customername))
        p.drawString(200, y_position, str(enquiry.contact))
        p.drawString(280, y_position, str(enquiry.get_status_display()))
        p.drawString(350, y_position, str(enquiry.products.name))
        p.drawString(450, y_position, str(enquiry.closuredate))
        p.drawString(550, y_position, str(enquiry.remarks))

        y_position -= 20

    p.showPage()
    p.save()

    buffer.seek(0)
    response.write(buffer.getvalue())
    return response


import openpyxl
from django.http import HttpResponse
from django.db.models import Q
from .models import Enquiry, User

def export_enquiries_xlsx(request):
    current_user = request.user
    selected_user_id = request.GET.get('user')
    selected_user = current_user

    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()
        # Managers/superusers see all enquiries (non-confirmed, non-lost, or reverted)
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        )
    else:
        # Other users see only their own created or assigned enquiries
        enquiries = Enquiry.objects.filter(
            Q(is_confirmed=False, is_lost=False) | Q(is_reverted=True)
        ).filter(
            Q(created_by=current_user) | Q(executive__name=current_user.username)
        )

    # Create in-memory workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Enquiries"
    
    # Add headers
    sheet.append(['ID', 'Company Name', 'Customer Name', 'Contact', 'Status', 'Policies', 'Enquiry Date', 'Remarks'])

    for enquiry in enquiries:
        sheet.append([
            enquiry.id,
            enquiry.companyname,
            enquiry.customername,
            enquiry.contact,
            enquiry.get_status_display(),
            enquiry.products.name,
            enquiry.closuredate,
            enquiry.remarks
        ])

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="enquiries.xlsx"'

    workbook.save(response)
    return response


from django.utils.timezone import now

# **Export Confirmed Orders Due for Renewal as CSV**
def export_confirmed_renewals_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="confirmed_renewals.csv"'
    
    writer = csv.writer(response)
    writer.writerow([ 'Company Name', 'Customer Name', 'Contact','policies','Start Date', 'End Date','VID', ])

    today = now().date()
    confirmed_orders_due_for_renewal = []

    confirmed_orders = Enquiry.objects.filter(is_confirmed=True)
    for order in confirmed_orders:
        for confirmed_entry in order.confirmed_enquiry_set.all():
            renewal_date = confirmed_entry.enddate - timedelta(days=confirmed_entry.renewal_days)
            if today >= renewal_date and not confirmed_entry.renewal_flag:
                confirmed_orders_due_for_renewal.append(confirmed_entry)

    for entry in confirmed_orders_due_for_renewal:
        writer.writerow([
            entry.enquiry.companyname, entry.enquiry.customername,
            entry.enquiry.contact,entry.enquiry.products.name,entry.startdate, entry.enddate, entry.vid
        ])

    return response


from openpyxl.utils import get_column_letter

def export_confirmed_renewals_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="confirmed_renewals.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Confirmed Renewals"

    columns = ['Company Name', 'Customer Name', 'Contact', 'Policies', 'Start Date', 'End Date', 'VID']
    
    # **Write column headers**
    for col_num, column_title in enumerate(columns, start=1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = column_title

    today = now().date()
    confirmed_orders_due_for_renewal = []

    confirmed_orders = Enquiry.objects.filter(is_confirmed=True)
    for order in confirmed_orders:
        for confirmed_entry in order.confirmed_enquiry_set.all():
            renewal_date = confirmed_entry.enddate - timedelta(days=confirmed_entry.renewal_days)
            if today >= renewal_date and not confirmed_entry.renewal_flag:
                confirmed_orders_due_for_renewal.append(confirmed_entry)

    row_num = 2  # Data starts from row 2
    for entry in confirmed_orders_due_for_renewal:
        ws[f"A{row_num}"] = entry.enquiry.companyname
        ws[f"B{row_num}"] = entry.enquiry.customername
        ws[f"C{row_num}"] = entry.enquiry.contact
        ws[f"D{row_num}"] = entry.enquiry.products.name if entry.enquiry.products else "N/A"
        ws[f"E{row_num}"] = entry.startdate.strftime('%Y-%m-%d')  # Format date properly
        ws[f"F{row_num}"] = entry.enddate.strftime('%Y-%m-%d')
        ws[f"G{row_num}"] = entry.vid
        row_num += 1

    wb.save(response)
    return response

def export_confirmed_renewals_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="confirmed_renewals.pdf"'

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y_position = height - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(200, y_position, "Confirmed Orders Due for Renewal")
    y_position -= 30

    pdf.setFont("Helvetica-Bold", 10)
    headers = ['Company Name', 'Customer Name', 'Contact', 'Policies', 'Start Date', 'End Date', 'VID']
    
    # **Draw table headers**
    x_positions = [30, 130, 210, 290, 390, 460, 530]  # Column positions
    for i, header in enumerate(headers):
        pdf.drawString(x_positions[i], y_position, header)
    y_position -= 15  # Move down after headers

    pdf.setFont("Helvetica", 10)
    today = now().date()
    confirmed_orders_due_for_renewal = []

    confirmed_orders = Enquiry.objects.filter(is_confirmed=True)
    for order in confirmed_orders:
        for confirmed_entry in order.confirmed_enquiry_set.all():
            renewal_date = confirmed_entry.enddate - timedelta(days=confirmed_entry.renewal_days)
            if today >= renewal_date and not confirmed_entry.renewal_flag:
                confirmed_orders_due_for_renewal.append(confirmed_entry)

    for entry in confirmed_orders_due_for_renewal:
        if y_position < 50:  # Start new page if space runs out
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y_position = height - 40

        pdf.drawString(x_positions[0], y_position, str(entry.enquiry.companyname))
        pdf.drawString(x_positions[1], y_position, str(entry.enquiry.customername))
        pdf.drawString(x_positions[2], y_position, str(entry.enquiry.contact))  # **Fix: Convert int to str**
        pdf.drawString(x_positions[3], y_position, str(entry.enquiry.products.name if entry.enquiry.products else "N/A"))
        pdf.drawString(x_positions[4], y_position, entry.startdate.strftime('%Y-%m-%d'))
        pdf.drawString(x_positions[5], y_position, entry.enddate.strftime('%Y-%m-%d'))
        pdf.drawString(x_positions[6], y_position, str(entry.vid))  # **Fix: Convert VID to str**

        y_position -= 15  # Move down for next row

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    response.write(buffer.read())

    return response

def export_past_policies_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="past_policies.csv"'

    writer = csv.writer(response)
    writer.writerow(['Company Name', 'Customer Name', 'Contact', 'Policy Type', 'Start Date', 'End Date', 'VID'])

    past_policies = confirmed_enquiry.objects.filter(revert=True)

    for entry in past_policies:
        writer.writerow([
            entry.enquiry.companyname,
            entry.enquiry.customername,
            entry.enquiry.contact,
            entry.policy_type,
            entry.startdate.strftime('%Y-%m-%d'),
            entry.enddate.strftime('%Y-%m-%d'),
            entry.vid,
        ])

    return response


def export_past_policies_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="past_policies.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Past Policies"

    # **Define column headers**
    columns = ['Company Name', 'Customer Name', 'Contact', 'Policy Type', 'Start Date', 'End Date', 'VID']
    ws.append(columns)  # Adds header row

    # **Fetch data**
    past_policies = confirmed_enquiry.objects.filter(revert=True)

    # **Write data to rows**
    for entry in past_policies:
        ws.append([
            entry.enquiry.companyname,
            entry.enquiry.customername,
            str(entry.enquiry.contact),  # Convert to string
            entry.policy_type,
            entry.startdate.strftime('%Y-%m-%d'),
            entry.enddate.strftime('%Y-%m-%d'),
            str(entry.vid)  # Convert to string
        ])

    # **Save workbook to response**
    wb.save(response)
    return response


def export_past_policies_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="past_policies.pdf"'

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y_position = height - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(200, y_position, "Past Policies Report")
    y_position -= 30

    pdf.setFont("Helvetica-Bold", 10)
    headers = ['Company Name', 'Customer Name', 'Contact', 'Policy Type', 'Start Date', 'End Date', 'VID']
    
    # **Draw table headers**
    x_positions = [30, 130, 210, 290, 390, 460, 530]  # Column positions
    for i, header in enumerate(headers):
        pdf.drawString(x_positions[i], y_position, header)
    y_position -= 15  # Move down after headers

    pdf.setFont("Helvetica", 10)
    past_policies = confirmed_enquiry.objects.filter(revert=True)

    for entry in past_policies:
        if y_position < 50:  # Start new page if space runs out
            pdf.showPage()
            pdf.setFont("Helvetica", 10)
            y_position = height - 40

        pdf.drawString(x_positions[0], y_position, str(entry.enquiry.companyname))
        pdf.drawString(x_positions[1], y_position, str(entry.enquiry.customername))
        pdf.drawString(x_positions[2], y_position, str(entry.enquiry.contact))  # Convert to string
        pdf.drawString(x_positions[3], y_position, str(entry.policy_type))
        pdf.drawString(x_positions[4], y_position, entry.startdate.strftime('%Y-%m-%d'))
        pdf.drawString(x_positions[5], y_position, entry.enddate.strftime('%Y-%m-%d'))
        pdf.drawString(x_positions[6], y_position, str(entry.vid))  # Convert to string

        y_position -= 15  # Move down for next row

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    response.write(buffer.read())

    return response


def export_enq_home_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="enq_home.csv"'

    writer = csv.writer(response)
    writer.writerow(['Client Name', 'Policy No', 'Policy Type', 'Insurance Company', 'Insurance Branch', 'Start Date', 'End Date', 'Contact'])

    # Fetch all data
    retailpolicies_data = retailpolicies.objects.all().order_by('-id')
    personal_acc_data = personal_Accident.objects.all().order_by('-id')

    # Write Retail Policies Data
    for entry in retailpolicies_data:
        writer.writerow([
            entry.cid.cname if entry.cid else "N/A",  # Client Name
            entry.polno,  # Policy No
            entry.retpol.insname if entry.retpol else "N/A",  # Policy Type
            entry.inscomp.insname if entry.inscomp else "N/A",  # Insurance Company
            entry.insbranch.insbranch if entry.insbranch else "N/A",  # Insurance Branch
            entry.strtdate if entry.strtdate else "N/A",  # Start Date
            entry.enddate if entry.enddate else "N/A",  # End Date
            entry.cid.ccontact if entry.cid else "N/A"  # Contact
        ])

    # Write Personal Accident Policies Data
    for entry in personal_acc_data:
        writer.writerow([
            entry.cid.cname if entry.cid else "N/A",  # Client Name
            entry.policy_no,  # Policy No
            "Personal Accident",  # Policy Type (fixed value)
            entry.inscomp_name.insname if entry.inscomp_name else "N/A",  # Insurance Company
            entry.ins_branch.insbranch if entry.ins_branch else "N/A",  # Insurance Branch
            entry.start_date if entry.start_date else "N/A",  # Start Date
            entry.end_date if entry.end_date else "N/A",  # End Date
            entry.cid.ccontact if entry.cid else "N/A"  # Contact
        ])

    return response

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font

def export_enq_home_xlsx(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="enq_home.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiry Home"

    # Define column headers
    columns = ['Client Name', 'Policy No', 'Policy Type', 'Insurance Company', 'Insurance Branch', 'Start Date', 'End Date', 'Contact']

    # Apply bold styling to headers
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    # Fetch all data
    retailpolicies_data = retailpolicies.objects.all().order_by('-id')
    personal_acc_data = personal_Accident.objects.all().order_by('-id')

    row_num = 2  # Start from the second row (first row is header)

    # Write Retail Policies Data
    for entry in retailpolicies_data:
        ws.append([
            entry.cid.cname if entry.cid else "N/A",  # Client Name
            entry.polno,  # Policy No
            entry.retpol.insname if entry.retpol else "N/A",  # Policy Type
            entry.inscomp.insname if entry.inscomp else "N/A",  # Insurance Company
            entry.insbranch.insbranch if entry.insbranch else "N/A",  # Insurance Branch
            entry.strtdate.strftime("%Y-%m-%d") if entry.strtdate else "N/A",  # Start Date
            entry.enddate.strftime("%Y-%m-%d") if entry.enddate else "N/A",  # End Date
            entry.cid.ccontact if entry.cid else "N/A"  # Contact
        ])
        row_num += 1

    # Write Personal Accident Policies Data
    for entry in personal_acc_data:
        ws.append([
            entry.cid.cname if entry.cid else "N/A",  # Client Name
            entry.policy_no,  # Policy No
            "Personal Accident",  # Policy Type (fixed value)
            entry.inscomp_name.insname if entry.inscomp_name else "N/A",  # Insurance Company
            entry.ins_branch.insbranch if entry.ins_branch else "N/A",  # Insurance Branch
            entry.start_date.strftime("%Y-%m-%d") if entry.start_date else "N/A",  # Start Date
            entry.end_date.strftime("%Y-%m-%d") if entry.end_date else "N/A",  # End Date
            entry.cid.ccontact if entry.cid else "N/A"  # Contact
        ])
        row_num += 1

    # Save to response
    wb.save(response)
    return response

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO

def export_enq_home_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="enq_home.pdf"'

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50  # Start position

    # Title
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Active Policies")
    y -= 30

    # Define column headers and positions
    columns = [
        (30, "Client Name"),
        (110, "Policy No"),
        (210, "Policy Type"),
        (310, "Company"),
        (370, "Branch"),
        (420, "Start Date"),
        (480, "End Date"),
        (540, "Contact"),

    ]

    # Draw headers
    p.setFont("Helvetica-Bold", 10)
    for x, title in columns:
        p.drawString(x, y, title)
    y -= 20

    p.setFont("Helvetica", 10)

    # Fetch data from both models
    retailpolicies_data = retailpolicies.objects.all().order_by('-id')
    pa_policies_data = personal_Accident.objects.all().order_by('-id')

    # Function to write a row of data
    def write_row(entry, policy_type):
        nonlocal y
        p.drawString(30, y, entry.cid.cname if entry.cid else "N/A")  # Client Name
        
        # Get Policy Number
        policy_no = entry.polno if hasattr(entry, "polno") else entry.policy_no if hasattr(entry, "policy_no") else "N/A"
        p.drawString(110, y, policy_no)

        p.drawString(210, y, policy_type)  # Policy Type

        # Insurance Company & Branch
        company = entry.inscomp.insname if hasattr(entry, "inscomp") else entry.inscomp_name.insname if hasattr(entry, "inscomp_name") else "N/A"
        p.drawString(310, y, company)

        branch = entry.insbranch.insbranch if hasattr(entry, "insbranch") else entry.ins_branch.insbranch if hasattr(entry, "ins_branch") else "N/A"
        p.drawString(370, y, branch)

        # Start Date & End Date
        start_date = entry.strtdate.strftime("%Y-%m-%d") if hasattr(entry, "strtdate") else entry.start_date.strftime("%Y-%m-%d") if hasattr(entry, "start_date") else "N/A"
        p.drawString(420, y, start_date)

        end_date = entry.enddate.strftime("%Y-%m-%d") if hasattr(entry, "enddate") else entry.end_date.strftime("%Y-%m-%d") if hasattr(entry, "end_date") else "N/A"
        p.drawString(480, y, end_date)

        # Contact Person
        contact = str(entry.cid.ccontact) if entry.cid and hasattr(entry.cid, "ccontact") else "N/A"
        p.drawString(540, y, contact)



        y -= 20  # Move to next row

        # Page Break Condition
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50

    # Write Retail Policies
    for entry in retailpolicies_data:
        write_row(entry, "Retail Policy")

    # Write Personal Accident Policies
    for entry in pa_policies_data:
        write_row(entry, "Personal Accident")

    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def i_renewals_export_csv(request):
    today = now().date()
    renewal_date = today + timedelta(days=30)

    renewal_policies = retailpolicies.objects.filter(enddate__lte=renewal_date, enddate__gte=today).order_by('-id')
    personal_acc = personal_Accident.objects.filter(end_date__lte=renewal_date, end_date__gte=today).order_by('-id')

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="Retail||_renewals.csv"'
    
    writer = csv.writer(response)
    writer.writerow(["Client Name", "Policy No", "Policy Type", "Insurance Company", "Branch", "Start Date", "End Date", "Contact"])
    
    for policy in renewal_policies:
        writer.writerow([
            policy.cid.cname if policy.cid else "N/A",
            policy.polno,
            policy.retpol.insname,
            policy.inscomp.insname if policy.inscomp else "N/A",
            policy.insbranch.insbranch if policy.insbranch else "N/A",
            policy.strtdate.strftime("%Y-%m-%d") if policy.strtdate else "N/A",
            policy.enddate.strftime("%Y-%m-%d") if policy.enddate else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ])

    for policy in personal_acc:
        writer.writerow([
            policy.cid.cname if policy.cid else "N/A",
            policy.policy_no,
            policy.category.insname,
            policy.inscomp_name.insname if policy.inscomp_name else "N/A",
            policy.ins_branch.insbranch if policy.ins_branch else "N/A",
            policy.start_date.strftime("%Y-%m-%d") if policy.start_date else "N/A",
            policy.end_date.strftime("%Y-%m-%d") if policy.end_date else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ])

    return response


def i_renewals_export_xlsx(request):
    today = now().date()
    renewal_date = today + timedelta(days=30)

    renewal_policies = retailpolicies.objects.filter(enddate__lte=renewal_date, enddate__gte=today).order_by('-id')
    personal_acc = personal_Accident.objects.filter(end_date__lte=renewal_date, end_date__gte=today).order_by('-id')

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Retail_renewals.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Client Name", "Policy No", "Policy Type", "Insurance Company", "Branch", "Start Date", "End Date", "Contact"])
    
    for policy in renewal_policies:
        ws.append([
            policy.cid.cname if policy.cid else "N/A",
            policy.polno,
            policy.retpol.insname,
            policy.inscomp.insname if policy.inscomp else "N/A",
            policy.insbranch.insbranch if policy.insbranch else "N/A",
            policy.strtdate.strftime("%Y-%m-%d") if policy.strtdate else "N/A",
            policy.enddate.strftime("%Y-%m-%d") if policy.enddate else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ])

    for policy in personal_acc:
        ws.append([
            policy.cid.cname if policy.cid else "N/A",
            policy.policy_no,
            policy.category.insname,
            policy.inscomp_name.insname if policy.inscomp_name else "N/A",
            policy.ins_branch.insbranch if policy.ins_branch else "N/A",
            policy.start_date.strftime("%Y-%m-%d") if policy.start_date else "N/A",
            policy.end_date.strftime("%Y-%m-%d") if policy.end_date else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ])

    wb.save(response)
    return response


def i_renewals_export_pdf(request):
    today = now().date()
    renewal_date = today + timedelta(days=30)

    renewal_policies = retailpolicies.objects.filter(enddate__lte=renewal_date, enddate__gte=today).order_by('-id')
    personal_acc = personal_Accident.objects.filter(end_date__lte=renewal_date, end_date__gte=today).order_by('-id')

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="renewals.pdf"'

    p = canvas.Canvas(response)
    y = 800  # Start position for writing data

    # Add title
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Renewal Policies Report")
    y -= 30

    # Table Headers
    p.setFont("Helvetica-Bold", 10)
    headers = ["Client Name", "Policy No", "Policy Type", "Company", "Branch", "Start Date", "End Date", "Contact"]
    x_positions = [20, 100, 200, 300, 360, 410, 470, 530]




    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)
    
    y -= 20  # Move to next line

    # Table Rows
    p.setFont("Helvetica", 10)
    for policy in renewal_policies:
        row = [
            policy.cid.cname if policy.cid else "N/A",
            policy.polno,
            "Retail Policy",
            policy.inscomp.insname if policy.inscomp else "N/A",
            policy.insbranch.insbranch if policy.insbranch else "N/A",
            policy.strtdate.strftime("%Y-%m-%d") if policy.strtdate else "N/A",
            policy.enddate.strftime("%Y-%m-%d") if policy.enddate else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ]
        for i, value in enumerate(row):
            p.drawString(x_positions[i], y, str(value))
        y -= 20

    for policy in personal_acc:
        row = [
            policy.cid.cname if policy.cid else "N/A",
            policy.policy_no,
            "Personal Accident",
            policy.inscomp_name.insname if policy.inscomp_name else "N/A",
            policy.ins_branch.insbranch if policy.ins_branch else "N/A",
            policy.start_date.strftime("%Y-%m-%d") if policy.start_date else "N/A",
            policy.end_date.strftime("%Y-%m-%d") if policy.end_date else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ]
        for i, value in enumerate(row):
            p.drawString(x_positions[i], y, str(value))
        y -= 20

    p.showPage()
    p.save()
    
    return response



def export_retail_past_policy_csv(request):
    """Export all retail past policies as CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=retail_past_policies.csv'

    writer = csv.writer(response)
    writer.writerow(['Client Name', 'Policy No', 'Category', 'Insurer Name', 'Branch Name',  'Start Date', 'End Date', 'Contact'])

    past_policies = retailpolicies.objects.filter(renewals=True)
    personal = personal_Accident.objects.filter(renewals=True)

    for policy in past_policies:
        writer.writerow([policy.cid.cname, policy.polno, policy.retpol.insname, policy.inscomp.insname,
                         policy.insbranch.insbranch, policy.strtdate,
                         policy.enddate, policy.cid.ccontact])

    for policy in personal:
        writer.writerow([policy.cid.cname, policy.policy_no, policy.category.insname, policy.inscomp_name.insname,
                         policy.ins_branch.insbranch,  policy.start_date,
                         policy.end_date, policy.cid.ccontact])

    return response

def export_retail_past_policy_xlsx(request):
    """Export all retail past policies as Excel (XLSX) using openpyxl."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=retail_past_policies.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Past Policies"

    # Define column headers
    columns = ['Client Name', 'Policy No', 'Category', 'Insurer Name', 'Branch Name', 'Start Date', 'End Date', 'Contact']
    ws.append(columns)  # Append column headers to first row

    # Fetch data
    past_policies = retailpolicies.objects.filter(renewals=True)
    personal = personal_Accident.objects.filter(renewals=True)

    # Write data for retail policies
    for policy in past_policies:
        ws.append([
            policy.cid.cname,
            policy.polno,
            policy.retpol.insname,
            policy.inscomp.insname,
            policy.insbranch.insbranch,
            policy.strtdate.strftime('%Y-%m-%d'),
            policy.enddate.strftime('%Y-%m-%d'),
            policy.cid.ccontact
        ])

    # Write data for personal accident policies
    for policy in personal:
        ws.append([
            policy.cid.cname,
            policy.policy_no,
            policy.category.insname,
            policy.inscomp_name.insname,
            policy.ins_branch.insbranch,
            policy.start_date.strftime('%Y-%m-%d'),
            policy.end_date.strftime('%Y-%m-%d'),
            policy.cid.ccontact
        ])

    # Auto-adjust column widths
    for col_num, column_title in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_num)].auto_size = True

    # Save workbook to response
    wb.save(response)
    return response



def export_retail_past_policy_pdf(request):
    """Export all retail past policies as a structured PDF."""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="retail_past_policies.pdf"'

    p = canvas.Canvas(response)
    y = 800  # Start position for text

    # Title
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Retail Past Policies Report")
    y -= 30

    # Table Headers
    p.setFont("Helvetica-Bold", 10)
    headers = ["Client Name", "Policy No", "Category", "Insurer", "Branch", "Start Date", "End Date", "Contact"]
    x_positions =  [20, 100, 200, 300, 360, 410, 470, 530]


    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)

    y -= 20  # Move to next row

    # Fetch policies
    past_policies = retailpolicies.objects.filter(renewals=True)
    personal = personal_Accident.objects.filter(renewals=True)

    # Table Rows (Retail Policies)
    p.setFont("Helvetica", 10)
    for policy in past_policies:
        row = [
            policy.cid.cname if policy.cid else "N/A",
            policy.polno,
            policy.retpol.insname if policy.retpol else "N/A",
            policy.inscomp.insname if policy.inscomp else "N/A",
            policy.insbranch.insbranch if policy.insbranch else "N/A",
            policy.strtdate.strftime("%Y-%m-%d") if policy.strtdate else "N/A",
            policy.enddate.strftime("%Y-%m-%d") if policy.enddate else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ]
        for i, value in enumerate(row):
            p.drawString(x_positions[i], y, str(value))
        y -= 20

    # Table Rows (Personal Accident Policies)
    for policy in personal:
        row = [
            policy.cid.cname if policy.cid else "N/A",
            policy.policy_no,
            policy.category.insname if policy.category else "N/A",
            policy.inscomp_name.insname if policy.inscomp_name else "N/A",
            policy.ins_branch.insbranch if policy.ins_branch else "N/A",
            policy.start_date.strftime("%Y-%m-%d") if policy.start_date else "N/A",
            policy.end_date.strftime("%Y-%m-%d") if policy.end_date else "N/A",
            policy.cid.ccontact if policy.cid and hasattr(policy.cid, "ccontact") else "N/A",
        ]
        for i, value in enumerate(row):
            p.drawString(x_positions[i], y, str(value))
        y -= 20

    p.showPage()
    p.save()
    
    return response


from django.template.loader import get_template
from xhtml2pdf import pisa

def export_followup_csv(request):
    followups = FollowUp_History.objects.all().order_by('-id')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="followups.csv"'

    writer = csv.writer(response)
    writer.writerow(['Company Name', 'Follow-Up Name', 'Date', 'Time', 'Order Status'])

     # Write the CSV data
    for followup in followups:
        # Check if it's a Confirmed Order
        if followup.confirmed_order:
            # Combine Company Name with VID like "Logitech - VID-2025-03001"
            company_name = f"{followup.confirmed_order.enquiry.companyname} - {followup.confirmed_order.vid}"
            order_status = "Confirmed Order"
        else:
            # For Enquiry, just display the Company Name without VID
            company_name = followup.enquiry.companyname
            order_status = "Enquiry"
        
        # Write the row data
        writer.writerow([
            company_name,
            followup.foname,
            followup.fodate,
            followup.fotime,
            order_status
        ])

    return response


def export_followup_excel(request):
    followups = FollowUp_History.objects.all().order_by('-id')
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="followups.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('FollowUps')

    # Write header row
    columns = ['Company Name', 'Follow-Up Name', 'Date', 'Time', 'Order Status']
    for col_num, col_name in enumerate(columns):
        ws.write(0, col_num, col_name)

    # Write data rows
    row_num = 1
    for followup in followups:
        # Check if it's a Confirmed Order
        if followup.confirmed_order:
            # Combine Company Name with VID like "Logitech - VID-2025-03001"
            company_name = f"{followup.confirmed_order.enquiry.companyname} - {followup.confirmed_order.vid}"
            order_status = "Confirmed Order"
        else:
            # For Enquiry, just display the Company Name without VID
            company_name = followup.enquiry.companyname
            order_status = "Enquiry"
        ws.write(row_num, 0, company_name)
        ws.write(row_num, 1, followup.foname)
        ws.write(row_num, 2, str(followup.fodate))
        ws.write(row_num, 3, str(followup.fotime))
        ws.write(row_num, 4, order_status)
        row_num += 1

    wb.save(response)
    return response

import csv
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from .models import FollowUp_History

def export_followup_pdf(request):
    """Export all follow-ups as a structured PDF."""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="followups.pdf"'

    # Initialize the PDF canvas
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    y = height - 50  # Start position for text

    # Title
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Follow-Up Report")
    y -= 30

    # Table Headers
    p.setFont("Helvetica-Bold", 10)
    headers = ["Company Name", "Follow-Up Name", "Date", "Time", "Order Status"]
    x_positions = [20, 180, 300, 400, 490]

    # Draw the headers
    for i, header in enumerate(headers):
        p.drawString(x_positions[i], y, header)

    y -= 20  # Move to next row

    # Fetch Follow-Ups
    followups = FollowUp_History.objects.all().order_by('-id')
    styles = getSampleStyleSheet()
    wrap_width = 100

    # Table Rows
    p.setFont("Helvetica", 10)
    for followup in followups:
        # Combine Company Name with VID if confirmed order exists
        if followup.confirmed_order:
            company_name = f"{followup.confirmed_order.enquiry.companyname} - {followup.confirmed_order.vid}"
            order_status = "Confirmed Order"
        else:
            company_name = followup.enquiry.companyname
            order_status = "Enquiry"

        # Wrap text for Company Name and Follow-Up Name
        company_name_paragraph = Paragraph(company_name, styles['Normal'])
        followup_name_paragraph = Paragraph(followup.foname, styles['Normal'])

        # --- ✅ Draw Company Name (wrapped text)
        company_name_paragraph.wrapOn(p, 150, 40)
        company_name_paragraph.drawOn(p, 20, y)

        # --- ✅ Draw Follow-Up Name (wrapped text)
        followup_name_paragraph.wrapOn(p, 110, 40)
        followup_name_paragraph.drawOn(p, 180, y)

        # Calculate the maximum height after wrapping
        max_height = max(company_name_paragraph.height, followup_name_paragraph.height)

        # --- ✅ Draw Date, Time, and Status properly aligned
        p.drawString(300, y, followup.fodate.strftime("%Y-%m-%d") if followup.fodate else "N/A")
        p.drawString(400, y, followup.fotime.strftime("%I:%M %p") if followup.fotime else "N/A")
        p.drawString(490, y, order_status)

        # --- ✅ Move y-axis according to max height to avoid overlap
        y -= max_height + 10

        # Avoid running off the page
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = height - 50

    # Finalize the PDF
    p.showPage()
    p.save()
    return response





def export_retail_followup_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="retail_followup_history.csv"'

    writer = csv.writer(response)
    writer.writerow(['Client Name - Policy Number', 'Follow Up Name', 'Follow Up Date', 'Follow Up Time'])

    followups = Retail_FollowUp_History.objects.all().order_by('-id')

    for followup in followups:
        # Check if it's a Retail Policy
        if followup.Retail:
            client_info = f"{followup.Retail.cid.cname} - {followup.Retail.polno}"
        # Check if it's a Personal Accident Policy
        elif followup.Personal:
            client_info = f"{followup.Personal.cid.cname} - {followup.Personal.policy_no}"
        # If no policy exists, show N/A
        else:
            client_info = "N/A - N/A"

        # Write data to CSV file
        writer.writerow([
            client_info,
            followup.foname,
            followup.fodate,
            followup.fotime,
        ])

    return response

import xlwt
from django.http import HttpResponse
from .models import Retail_FollowUp_History

def export_retail_followup_xlsx(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="retail_followup_history.xls"'

    workbook = xlwt.Workbook(encoding='utf-8')
    worksheet = workbook.add_sheet('Follow Up History')

    # Style for headers
    header_style = xlwt.easyxf('font: bold 1, color black; align: horiz center')

    # Sheet headers
    columns = ['Client Name', 'Follow Up Name', 'Follow Up Date', 'Follow Up Time', ]
    for col_num in range(len(columns)):
        worksheet.write(0, col_num, columns[col_num], header_style)
        worksheet.col(col_num).width = 6000  # Adjust column width

    # Fetch all follow-ups
    followups = Retail_FollowUp_History.objects.all().order_by('-id')
    row_num = 1

    # Loop through each follow-up and add it to the sheet
    for followup in followups:
        # Handle Client Name + Policy Number
        if followup.Retail:
            client_info = f"{followup.Retail.cid.cname} - {followup.Retail.polno}"
        elif followup.Personal:
            client_info = f"{followup.Personal.cid.cname} - {followup.Personal.policy_no}"
        else:
            client_info = "N/A - N/A"

        # Write each cell value
        worksheet.write(row_num, 0, client_info)
        worksheet.write(row_num, 1, followup.foname)
        worksheet.write(row_num, 2, str(followup.fodate))
        worksheet.write(row_num, 3, str(followup.fotime))
        
        row_num += 1

    # Save the workbook
    workbook.save(response)
    return response


from reportlab.lib.utils import simpleSplit

def export_retail_followup_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="retail_followup_history.pdf"'

    # Set up canvas
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, height - 30, "Follow Up History Report")
    p.setFont("Helvetica", 10)

    # Headers
    y = height - 60
    p.setFont("Helvetica-Bold", 10)
    p.drawString(30, y, "Client Name - Policy Number")
    p.drawString(220, y, "Follow Up Name")
    p.drawString(420, y, "Follow Up Date")
    p.drawString(500, y, "Follow Up Time")
    y -= 20

    # Data
    p.setFont("Helvetica", 10)
    followups = Retail_FollowUp_History.objects.all().order_by('-id')

    for followup in followups:
        # Handle Client Name + Policy Number
        if followup.Retail:
            client_info = f"{followup.Retail.cid.cname} - {followup.Retail.polno}"
        elif followup.Personal:
            client_info = f"{followup.Personal.cid.cname} - {followup.Personal.policy_no}"
        else:
            client_info = "N/A - N/A"

        # Wrap text if too long
        followup_name = followup.foname
        client_name_lines = simpleSplit(client_info, "Helvetica", 10, 180)
        followup_name_lines = simpleSplit(followup_name, "Helvetica", 10, 180)

        # Auto-wrap client name if it's too long
        max_lines = max(len(client_name_lines), len(followup_name_lines))
        for i in range(max_lines):
            if i < len(client_name_lines):
                p.drawString(30, y, client_name_lines[i])
            if i < len(followup_name_lines):
                p.drawString(220, y, followup_name_lines[i])

            # Only draw date/time on the first line to avoid overlap
            if i == 0:
                p.drawString(420, y, str(followup.fodate))
                p.drawString(500, y, str(followup.fotime))

            y -= 15

            # Check if the page is filled and create a new page
            if y <= 50:
                p.showPage()
                p.setFont("Helvetica", 10)
                y = height - 60
                p.drawString(30, y, "Client Name - Policy Number")
                p.drawString(220, y, "Follow Up Name")
                p.drawString(420, y, "Follow Up Date")
                p.drawString(500, y, "Follow Up Time")
                y -= 20

    # Finalize and save the PDF
    p.showPage()
    p.save()
    return response